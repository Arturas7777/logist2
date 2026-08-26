"""Отрисовка выгруженного XLSX в PNG — проверка, что бланк собрался правильно.

Читает docs/cmr_blank.xlsx и рисует его так, как его увидит Excel/Таблицы:
ширины столбцов и высоты строк переводятся обратно в миллиметры, рамки берутся
из стилей ячеек. Рядом кладётся HTML-рендер для сверки.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

OUT = Path(".cache")
XLSX = Path("docs") / "cmr_blank.xlsx"
FONT = Path(r"C:\Windows\Fonts\arial.ttf")
FONT_B = Path(r"C:\Windows\Fonts\arialbd.ttf")

PPMM = 1390 / 184.2  # как в cmr_frame.png
LINE_PX = {"thin": 1, "medium": 2, "thick": 3}
_cache: dict[tuple[bool, int], ImageFont.FreeTypeFont] = {}


def font(pt: float, bold: bool) -> ImageFont.FreeTypeFont:
    px = max(4, round(pt / 72 * 25.4 * PPMM))
    key = (bold, px)
    if key not in _cache:
        _cache[key] = ImageFont.truetype(str(FONT_B if bold else FONT), px)
    return _cache[key]


def main() -> None:
    ws = load_workbook(XLSX).active
    widths, heights = [], []
    for c in range(1, ws.max_column + 1):
        w = ws.column_dimensions[get_column_letter(c)].width or 8.43
        widths.append((w * 7 + 5) / (96 / 25.4))
    for r in range(1, ws.max_row + 1):
        heights.append((ws.row_dimensions[r].height or 15) * 25.4 / 72)

    xs = [0.0]
    for w in widths:
        xs.append(xs[-1] + w)
    ys = [0.0]
    for h in heights:
        ys.append(ys[-1] + h)
    print(f"sheet {xs[-1]:.1f} x {ys[-1]:.1f} mm, {len(widths)} cols, {len(heights)} rows")

    img = Image.new("RGB", (round(xs[-1] * PPMM), round(ys[-1] * PPMM)), "white")
    d = ImageDraw.Draw(img)
    at = lambda v: round(v * PPMM)  # noqa: E731

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            x0, x1, y0, y1 = at(xs[c - 1]), at(xs[c]), at(ys[r - 1]), at(ys[r])
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb == "FFF4F8FD":
                d.rectangle((x0, y0, x1, y1), fill="#eef4fc")
            b = cell.border
            for side, box in (
                (b.top, (x0, y0, x1, y0)), (b.bottom, (x0, y1, x1, y1)),
                (b.left, (x0, y0, x0, y1)), (b.right, (x1, y0, x1, y1)),
            ):
                if side and side.style:
                    d.line(box, fill="black", width=LINE_PX.get(side.style, 1))

    merged = {}
    for rng in ws.merged_cells.ranges:
        merged[(rng.min_row, rng.min_col)] = (rng.max_row, rng.max_col)

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            if not cell.value:
                continue
            r1, c1 = merged.get((r, c), (r, c))
            f = font(cell.font.size or 6, bool(cell.font.bold))
            lines = str(cell.value).split("\n")
            step = f.size * 1.15
            top = at(ys[r - 1])
            if (cell.alignment.vertical or "bottom") == "center":
                top += max(0, ((at(ys[r1]) - at(ys[r - 1])) - step * len(lines)) / 2)
            for i, line in enumerate(lines):
                anchor_x = at(xs[c - 1]) + 1
                align = cell.alignment.horizontal or "left"
                if align == "center":
                    anchor_x = (at(xs[c - 1]) + at(xs[c1])) / 2 - d.textlength(line, f) / 2
                elif align == "right":
                    anchor_x = at(xs[c1]) - 1 - d.textlength(line, f)
                d.text((anchor_x, top + i * step), line, font=f, fill="black")

    img.save(OUT / "cmr_xlsx.png")
    ref = Image.open(OUT / "cmr_frame.png").convert("RGB")
    side = Image.new("RGB", (ref.width + img.width + 8, max(ref.height, img.height)), "#888")
    side.paste(ref, (0, 0))
    side.paste(img, (ref.width + 8, 0))
    side.save(OUT / "cmr_xlsx_compare.png")
    print(f"png -> {OUT / 'cmr_xlsx.png'}\ncompare -> {OUT / 'cmr_xlsx_compare.png'}")


if __name__ == "__main__":
    main()

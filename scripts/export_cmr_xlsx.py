"""Выгрузка бланка CMR из шаблона в XLSX для доводки в Google Таблицах.

Сетка листа строится по реальным линиям бланка: каждая вертикальная линия даёт
границу столбца, каждая горизонтальная — границу строки. В Таблицах поэтому можно
тянуть любую границу, и она двигает ровно одну линию бланка, не ломая остальное.

Геометрия снимается с живого DOM через headless Chrome, то есть выгрузка всегда
соответствует ``templates/admin/cmr_editor.html``:

* линии — из границ элементов (``border-*``) и из шага линовки ``--pitch``;
* подписи — по строкам текста, каждая со своим кеглем и выключкой;
* поля ввода — прямоугольники ``input``/``textarea``, они подсвечены заливкой,
  а карта «имя поля → адрес ячейки» лежит на втором листе.

Запуск: ``python scripts\\export_cmr_xlsx.py``  (нужен openpyxl, dev-зависимость).
"""

from __future__ import annotations

import html
import json
import re
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / ".cache"
MEASURE_HTML = OUT / "cmr_measure.html"
XLSX = ROOT / "docs" / "cmr_blank.xlsx"
CHROME = Path(r"C:\Program Files\Google\Chrome\Application\chrome.exe")

sys.path.insert(0, str(ROOT / "scripts" / "debug"))

# Порог склейки координат: линии ближе этого расстояния считаем одной линией.
X_TOL, Y_TOL = 0.5, 0.35
# Ширину столбца Excel и Таблицы считают как (px - 5) / 7 знака, и ниже ~1.6 мм
# эта формула у них расходится. Поэтому столбцов уже MIN_COL_MM не создаём: близкие
# линии сливаются в одну. В бланке разделитель половин листа на разных полосах
# стоит с разбегом до 1.8 мм — в таблице он всё равно должен быть один столбец.
MIN_COL_MM = 1.7
MIN_ROW_MM = 0.6
# Толщина линии -> тип рамки Excel. hair не используем: он рисуется точками.
WEIGHTS = ((0.22, "thin"), (0.47, "medium"), (99.0, "thick"))
FIELD_FILL = PatternFill("solid", fgColor="FFF4F8FD")
NUMERAL = re.compile(r"^\d{1,2}$")

MEASURE_JS = r"""
<script>
window.addEventListener("load", function () {
    var frame = document.querySelector(".cmr-frame");
    var F = frame.getBoundingClientRect();
    var MM = 96 / 25.4;
    var mm = function (v) { return Math.round(v / MM * 1000) / 1000; };
    var out = {w: mm(F.width), h: mm(F.height), hl: [], vl: [], texts: [], fields: []};

    // --- линии: границы элементов ---------------------------------------
    var sides = [["Top", 1, 1], ["Bottom", 1, -1], ["Left", 0, 1], ["Right", 0, -1]];
    var all = [frame].concat(Array.prototype.slice.call(frame.querySelectorAll("*")));
    for (var i = 0; i < all.length; i++) {
        var el = all[i], cs = getComputedStyle(el);
        if (cs.display === "none" || cs.visibility === "hidden") continue;
        // Овал CMR и водяной знак — не линии сетки.
        if (cs.borderTopLeftRadius !== "0px") continue;
        var r = el.getBoundingClientRect();
        if (r.width < 0.2 || r.height < 0.2) continue;
        for (var s = 0; s < sides.length; s++) {
            var name = sides[s][0], horiz = sides[s][1], dir = sides[s][2];
            if (cs["border" + name + "Style"] === "none") continue;
            var w = parseFloat(cs["border" + name + "Width"]);
            if (!w) continue;
            var edge = name === "Top" ? r.top : name === "Bottom" ? r.bottom
                     : name === "Left" ? r.left : r.right;
            var p = edge + dir * w / 2;
            if (horiz) {
                out.hl.push([mm(p - F.top), mm(r.left - F.left), mm(r.right - F.left), mm(w)]);
            } else {
                out.vl.push([mm(p - F.left), mm(r.top - F.top), mm(r.bottom - F.top), mm(w)]);
            }
        }
    }

    // --- линии: линовка внутри textarea ---------------------------------
    var ruled = frame.querySelectorAll(".ruled");
    for (var i = 0; i < ruled.length; i++) {
        var el = ruled[i], r = el.getBoundingClientRect();
        var pitch = parseFloat(getComputedStyle(el).getPropertyValue("--pitch"));
        if (!pitch) continue;
        var h = mm(r.height), y0 = mm(r.top - F.top);
        for (var k = 1; k * pitch <= h + 0.05; k++) {
            out.hl.push([Math.round((y0 + k * pitch) * 1000) / 1000,
                         mm(r.left - F.left), mm(r.right - F.left), 0.14]);
        }
    }

    // --- тексты: по визуальным строкам ----------------------------------
    var walker = document.createTreeWalker(frame, NodeFilter.SHOW_TEXT);
    var node, rng = document.createRange();
    while ((node = walker.nextNode())) {
        var s = node.nodeValue;
        if (!s.trim()) continue;
        // Водяной знак — декорация, в таблицу не переносим.
        if (node.parentElement.closest(".watermark")) continue;
        var cs = getComputedStyle(node.parentElement);
        var lines = [];
        for (var i = 0; i < s.length; i++) {
            rng.setStart(node, i);
            rng.setEnd(node, i + 1);
            var r = rng.getBoundingClientRect();
            if (!r.width && !r.height) continue;
            var line = null;
            for (var j = 0; j < lines.length; j++) {
                if (Math.abs(lines[j].top - r.top) < 1.5) { line = lines[j]; break; }
            }
            if (!line) {
                line = {top: r.top, bottom: r.bottom, left: r.left, right: r.right, t: ""};
                lines.push(line);
            }
            line.left = Math.min(line.left, r.left);
            line.right = Math.max(line.right, r.right);
            line.bottom = Math.max(line.bottom, r.bottom);
            line.t += s[i];
        }
        for (var j = 0; j < lines.length; j++) {
            var L = lines[j], t = L.t.replace(/\s+/g, " ").trim();
            if (!t) continue;
            out.texts.push({
                t: t, x: mm(L.left - F.left), y: mm(L.top - F.top),
                w: mm(L.right - L.left), h: mm(L.bottom - L.top),
                pt: Math.round(parseFloat(cs.fontSize) * 0.75 * 10) / 10,
                b: parseInt(cs.fontWeight, 10) >= 600 ? 1 : 0,
                al: cs.textAlign
            });
        }
    }

    // --- поля ввода ------------------------------------------------------
    var fields = frame.querySelectorAll("input, textarea");
    for (var i = 0; i < fields.length; i++) {
        var el = fields[i], r = el.getBoundingClientRect(), cs = getComputedStyle(el);
        out.fields.push({
            name: el.name || "", x: mm(r.left - F.left), y: mm(r.top - F.top),
            w: mm(r.width), h: mm(r.height),
            pt: Math.round(parseFloat(cs.fontSize) * 0.75 * 10) / 10,
            al: cs.textAlign, multi: el.tagName === "TEXTAREA" ? 1 : 0
        });
    }

    var pre = document.createElement("pre");
    pre.id = "geom";
    pre.textContent = JSON.stringify(out);
    document.body.replaceChildren(pre);
});
</script>
"""


def measure() -> dict:
    """Снимает геометрию бланка из headless Chrome."""
    import render_cmr

    # preview_html, а не render_html: мерить надо печатную геометрию, а на
    # экране у бланка свои, более толстые линии.
    page = render_cmr.preview_html(blank=True).replace("</body>", MEASURE_JS + "</body>")
    OUT.mkdir(exist_ok=True)
    MEASURE_HTML.write_text(page, encoding="utf-8")
    dom = subprocess.run(
        [
            str(CHROME),
            "--headless=new",
            "--disable-gpu",
            "--window-size=900,1400",
            "--virtual-time-budget=4000",
            "--dump-dom",
            MEASURE_HTML.as_uri(),
        ],
        check=True,
        capture_output=True,
    ).stdout.decode("utf-8", "replace")
    body = dom.split('<pre id="geom">', 1)[1].split("</pre>", 1)[0]
    return json.loads(html.unescape(body))


def axis(values: list[float], size: float, tol: float, min_gap: float) -> list[float]:
    """Границы сетки по одной оси.

    Координаты линий сначала собираются в кластеры по одиночной связи (шаг ``tol``),
    затем самые близкие кластеры сливаются, пока между ними меньше ``min_gap``.
    Слияние взвешено по числу линий, поэтому граница садится на самую «плотную»
    линию, а не в середину случайной пары.
    """
    groups: list[list[float]] = []
    for v in sorted(values):
        if groups and v - groups[-1][-1] <= tol:
            groups[-1].append(v)
        else:
            groups.append([v])
    items = [(sum(g) / len(g), len(g)) for g in groups]

    while len(items) > 1:
        gap, i = min((items[k + 1][0] - items[k][0], k) for k in range(len(items) - 1))
        if gap >= min_gap:
            break
        (p1, n1), (p2, n2) = items[i], items[i + 1]
        items[i : i + 2] = [((p1 * n1 + p2 * n2) / (n1 + n2), n1 + n2)]

    grid = [p for p, _ in items if min_gap < p < size - min_gap]
    return [0.0, *grid, size]


def add_soft(grid: list[float], values: list[float], gap: float) -> list[float]:
    """Добавляет границы по краям текста там, где линий бланка нет.

    Без этого подписи, не разделённые линиями (например, номера граф 6-9), попали
    бы в одну ячейку и склеились. Кандидаты сначала группируются и сортируются по
    числу совпадений: край, с которого начинается несколько строк подписи, важнее
    случайного одиночного отступа, а места хватает не всем.
    """
    groups: list[list[float]] = []
    for v in sorted(values):
        if groups and v - groups[-1][-1] <= 0.4:
            groups[-1].append(v)
        else:
            groups.append([v])
    candidates = sorted(groups, key=lambda g: (-len(g), g[0]))

    out = list(grid)
    for g in candidates:
        v = sum(g) / len(g)
        if out[0] < v < out[-1] and all(abs(v - x) >= gap for x in out):
            out.append(v)
            out.sort()
    return out


def cell_of(grid: list[float], v: float) -> int:
    """Индекс полосы сетки, в которую попадает координата."""
    for i in range(len(grid) - 1):
        if v < grid[i + 1] - 0.01:
            return i
    return len(grid) - 2


def col_of(grid: list[float], x: float) -> int:
    """Столбец для левого края текста с притяжкой к близкой границе.

    Если подпись начинается почти на границе столбца, относим её к следующему —
    иначе номер графы и текст подписи попадают в одну ячейку и склеиваются.
    """
    i = cell_of(grid, x + 0.2)
    if i + 2 < len(grid) and grid[i + 1] - x < 1.0:
        return i + 1
    return i


def nearest(grid: list[float], v: float, tol: float) -> int | None:
    """Индекс ближайшей границы сетки, если она достаточно близко."""
    best = min(range(len(grid)), key=lambda i: abs(grid[i] - v))
    return best if abs(grid[best] - v) <= tol else None


def row_of(grid: list[float], y: float) -> int:
    """Строка для текстовой строки: ближайшая граница сетки к её верху.

    Границы строк как раз и ставятся по верху строк текста, поэтому притяжка к
    ближайшей границе точнее, чем поиск полосы, в которую попал центр: у мелких
    кеглей строчный бокс выше самой полосы и центр уезжает в соседнюю.
    """
    best = min(range(len(grid) - 1), key=lambda i: abs(grid[i] - y))
    return best


def weight(w: float) -> str:
    return next(name for limit, name in WEIGHTS if w <= limit)


def stronger(a: str | None, b: str) -> str:
    order = ["thin", "medium", "thick"]
    return b if a is None or order.index(b) > order.index(a) else a


def build(geom: dict) -> Workbook:
    xs = add_soft(
        axis([v[0] for v in geom["vl"]], geom["w"], X_TOL, MIN_COL_MM),
        [t["x"] for t in geom["texts"]] + [f["x"] for f in geom["fields"]],
        gap=MIN_COL_MM,
    )
    ys = add_soft(
        axis([h[0] for h in geom["hl"]], geom["h"], Y_TOL, MIN_ROW_MM),
        [t["y"] for t in geom["texts"]],
        gap=1.0,
    )
    ncol, nrow = len(xs) - 1, len(ys) - 1
    thin = [round(xs[c + 1] - xs[c], 2) for c in range(ncol) if xs[c + 1] - xs[c] < MIN_COL_MM - 0.01]
    if thin:
        print(f"warning: columns narrower than {MIN_COL_MM} mm: {thin}")

    wb = Workbook()
    wb._fonts[0].name, wb._fonts[0].sz = "Arial", 5.5
    ws = wb.active
    ws.title = "CMR"
    ws.sheet_view.showGridLines = False

    px = 96 / 25.4
    for c in range(ncol):
        ws.column_dimensions[get_column_letter(c + 1)].width = ((xs[c + 1] - xs[c]) * px - 5) / 7
    for r in range(nrow):
        ws.row_dimensions[r + 1].height = (ys[r + 1] - ys[r]) * 72 / 25.4

    # --- рамки ------------------------------------------------------------
    edges: dict[tuple[int, int], dict[str, str]] = {}

    def put(row: int, col: int, side: str, w: float) -> None:
        if not (0 <= row < nrow and 0 <= col < ncol):
            return
        box = edges.setdefault((row, col), {})
        box[side] = stronger(box.get(side), weight(w))

    for p, a, b, w in geom["hl"]:
        line = nearest(ys, p, MIN_ROW_MM)
        if line is None:
            continue
        c0, c1 = cell_of(xs, a + 0.05), cell_of(xs, b - 0.05)
        for c in range(c0, c1 + 1):
            if line < nrow:
                put(line, c, "top", w)
            if line > 0:
                put(line - 1, c, "bottom", w)
    for p, a, b, w in geom["vl"]:
        line = nearest(xs, p, MIN_COL_MM)
        if line is None:
            continue
        r0, r1 = cell_of(ys, a + 0.05), cell_of(ys, b - 0.05)
        for r in range(r0, r1 + 1):
            if line < ncol:
                put(r, line, "left", w)
            if line > 0:
                put(r, line - 1, "right", w)

    for (row, col), box in edges.items():
        ws.cell(row + 1, col + 1).border = Border(
            **{side: Side(style=style) for side, style in box.items()}
        )

    # --- подписи ----------------------------------------------------------
    cells: dict[tuple[int, int], dict] = {}
    spans: list[tuple[int, int, int]] = []
    for t in geom["texts"]:
        # Выключку по центру/правому краю сохраняем только если строка целиком
        # укладывается в свою клетку. Иначе (заголовок бланка, юридический текст —
        # они шире любого столбца) привязываем по левому краю, чтобы не уехали.
        align = {"center": "center", "right": "right"}.get(t["al"], "left")
        col = -1
        if align != "left":
            i = cell_of(xs, t["x"] + t["w"] / 2)
            if xs[i] - 0.4 <= t["x"] and t["x"] + t["w"] <= xs[i + 1] + 0.4:
                col = i
            else:
                align = "left"
        if col < 0:
            col = col_of(xs, t["x"])

        if NUMERAL.match(t["t"]) and t["b"] and t["pt"] >= 8:
            # Номер графы в бланке стоит вровень с двумя строками подписи, поэтому
            # в таблице он занимает те же строки — иначе съезжает на нижнюю.
            row = cell_of(ys, t["y"] + 0.15)
            last = cell_of(ys, t["y"] + t["h"] - 0.15)
            if last > row:
                spans.append((row, last, col))
        else:
            row = row_of(ys, t["y"])
        box = cells.setdefault(
            (row, col), {"lines": [], "pt": 0.0, "b": 0, "al": align, "xs": [], "w": 0.0}
        )
        box["lines"].append(t["t"])
        box["pt"] = max(box["pt"], t["pt"])
        box["b"] = max(box["b"], t["b"])
        box["xs"].append(t["x"])
        box["w"] = max(box["w"], t["w"])

    collisions = [
        (r, c, box["lines"])
        for (r, c), box in cells.items()
        if max(box["xs"]) - min(box["xs"]) > 0.6
    ]
    if collisions:
        print(f"warning: {len(collisions)} cells mix texts from different columns:")
        for r, c, lines in collisions:
            print(f"  {get_column_letter(c + 1)}{r + 1}: {' | '.join(lines)}")

    def room(row: int, col: int) -> float:
        """Ширина от левого края ячейки до следующей занятой в этой строке."""
        end = col
        while end + 1 < ncol and (row, end + 1) not in cells:
            end += 1
        return xs[end + 1] - xs[col]

    for (row, col), box in cells.items():
        cell = ws.cell(row + 1, col + 1)
        cell.value = "\n".join(box["lines"])
        size = max(4.0, round(box["pt"] * 2) / 2)
        # Excel и Таблицы обрезают текст, если справа занятая ячейка. У номеров
        # граф 19 и 20 подпись начинается почти вплотную, поэтому кегль поджимаем.
        free = room(row, col)
        if len(box["lines"]) == 1 and box["w"] > free - 0.4 and NUMERAL.match(box["lines"][0]):
            size = max(6.0, round(size * (free - 0.4) / box["w"] * 2) / 2)
        cell.font = Font(name="Arial", size=size, bold=bool(box["b"]))
        # По верху, а не по центру: последняя текстовая строка блока попадает в
        # полосу, растянутую до следующей линии бланка, и центрирование уронило бы
        # подпись на несколько миллиметров вниз.
        cell.alignment = Alignment(
            horizontal=box["al"], vertical="top", wrap_text=len(box["lines"]) > 1
        )

    for r0, r1, col in spans:
        if any((r, col) in cells for r in range(r0 + 1, r1 + 1)):
            continue
        ws.merge_cells(start_row=r0 + 1, end_row=r1 + 1, start_column=col + 1, end_column=col + 1)

    # --- поля ввода -------------------------------------------------------
    fields = wb.create_sheet("Поля")
    fields.append(["Поле в шаблоне", "Ячейка", "Строк"])
    for f in geom["fields"]:
        if not f["name"]:
            continue
        # У textarea заполняется первая линованная строка, у input — вся высота.
        row = cell_of(ys, f["y"] + (1.8 if f["multi"] else f["h"] / 2))
        col = cell_of(xs, f["x"] + 0.2)
        cell = ws.cell(row + 1, col + 1)
        if cell.value is None:
            cell.fill = FIELD_FILL
            cell.font = Font(name="Arial", size=max(4.0, round(f["pt"] * 2) / 2))
            cell.alignment = Alignment(
                horizontal={"center": "center", "right": "right"}.get(f["al"], "left"),
                vertical="center",
            )
        fields.append(
            [f["name"], f"{get_column_letter(col + 1)}{row + 1}", "много" if f["multi"] else ""]
        )
    fields.column_dimensions["A"].width = 26
    fields.column_dimensions["B"].width = 10
    fields.column_dimensions["C"].width = 8

    guide = wb.create_sheet("Как править")
    guide.column_dimensions["A"].width = 110
    for line in [
        "Бланк CMR — лист «CMR». Выгружен из шаблона logist2 (templates/admin/cmr_editor.html).",
        "",
        f"Габарит бланка: {geom['w']:.1f} x {geom['h']:.1f} мм, {ncol} столбцов x {nrow} строк, A4 с полями "
        f"{(210 - geom['w']) / 2:.1f} мм по бокам и {(297 - geom['h']) / 2:.1f} мм сверху/снизу.",
        "",
        "Как устроена сетка",
        "  • Каждая линия бланка — это граница столбца или строки. Тянете границу — двигается ровно одна линия.",
        "  • Толщина линий задана рамками ячеек: тонкая / средняя / толстая = волосяная / основная / жирная в бланке.",
        "  • Голубая заливка — места ввода. Карта «поле шаблона → ячейка» на листе «Поля».",
        "",
        "Пересчёт размеров",
        "  • Высота строки: 1 мм = 2.835 пункта (в Google Таблицах высота задаётся в пикселях: 1 мм = 3.78 px).",
        "  • Ширина столбца: 1 мм = 3.78 px. В Excel ширина в знаках: знаки = (px − 5) / 7.",
        "  • Суммы по всем столбцам и строкам должны остаться равны габариту выше, иначе бланк уедет с листа.",
        "",
        "Чего в таблице нет",
        "  • Водяного знака «CMR» на весь бланк — это фон, в таблице он не нужен.",
        "  • Овала вокруг «CMR» в шапке: в Таблицах его проще добавить рисунком поверх ячеек.",
        "",
        "Файл пересобирается командой: python scripts/export_cmr_xlsx.py",
    ]:
        guide.append([line])

    # --- печать -----------------------------------------------------------
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins = PageMargins(
        left=(210 - geom["w"]) / 2 / 25.4, right=(210 - geom["w"]) / 2 / 25.4,
        top=(297 - geom["h"]) / 2 / 25.4, bottom=(297 - geom["h"]) / 2 / 25.4,
        header=0, footer=0,
    )
    ws.print_area = f"A1:{get_column_letter(ncol)}{nrow}"
    print(f"grid: {ncol} columns x {nrow} rows, frame {geom['w']:.1f} x {geom['h']:.1f} mm")
    return wb


def main() -> None:
    geom = measure()
    print(f"lines: {len(geom['hl'])} horizontal, {len(geom['vl'])} vertical; "
          f"texts: {len(geom['texts'])}, fields: {len(geom['fields'])}")
    build(geom).save(XLSX)
    print(f"xlsx -> {XLSX}")


if __name__ == "__main__":
    main()
